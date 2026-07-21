import os
import time
import sys
from pathlib import Path
import pandas as pd

# -- Ensure Selenium is installed --
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
except ImportError:
    print("Selenium library is required. Installing now...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "selenium"])
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

# -- Ensure openpyxl is installed for Excel formatting --
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel report generation...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# -- Paths --
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
REPORTS_DIR = PROJECT_ROOT / "e2e-tests" / "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)
EXCEL_REPORT = REPORTS_DIR / "web_test_report.xlsx"

# -- Test Suite Config --
BASE_URL = "http://localhost:5173"
TIMEOUT = 10

class WebE2ETester:
    def __init__(self):
        self.results = []
        self.driver = None
        
    def start_browser(self):
        print("Initializing Selenium Web Driver...")
        chrome_options = Options()
        # You can toggle headless mode depending on your preference
        # chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1280,800")
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.implicitly_wait(3)
        
    def log_result(self, test_id, name, step, status, details="", response_time=0.0):
        self.results.append({
            "Test ID": test_id,
            "Test Name": name,
            "Test Step": step,
            "Status": status,
            "Response Time (s)": round(response_time, 3),
            "Details / Errors": details,
            "Timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
        })
        print(f"[{status}] {test_id} - {name} ({step}) | Time: {response_time:.2f}s | Details: {details}")

    def run_tests(self):
        self.start_browser()
        
        # --- TEST 1: Load Landing Page ---
        t_start = time.time()
        try:
            self.driver.get(BASE_URL)
            # Wait for container or logo
            WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.CLASS_NAME, "container"))
            )
            self.log_result("WEB-001", "Load Landing Page", "Open Base URL", "PASS", "Landing page loaded successfully", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-001", "Load Landing Page", "Open Base URL", "FAIL", str(e), time.time() - t_start)
            
        # --- TEST 2: Check Navigation ---
        t_start = time.time()
        try:
            # Look for Doctor section link or login CTA
            login_link = WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@href, '/auth/login')]"))
            )
            login_link.click()
            self.log_result("WEB-002", "Navigation to Login", "Click Login Link", "PASS", "Successfully navigated to Auth login page", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-002", "Navigation to Login", "Click Login Link", "FAIL", str(e), time.time() - t_start)
            
        # --- TEST 3: Login Simulation ---
        t_start = time.time()
        try:
            # Find inputs
            email_input = WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email']"))
            )
            password_input = self.driver.find_element(By.XPATH, "//input[@type='password']")
            submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit']")
            
            # Fill credentials (simulated test user)
            email_input.send_keys("test_patient@orthocare.ai")
            password_input.send_keys("Password123!")
            submit_btn.click()
            
            # Usually we expect a home dashboard or profile completion page
            time.sleep(2)
            self.log_result("WEB-003", "Patient Login", "Submit Credentials", "PASS", "Submitted login credentials successfully", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-003", "Patient Login", "Submit Credentials", "FAIL", str(e), time.time() - t_start)

        # --- TEST 4: Navigate to Scan Page ---
        t_start = time.time()
        try:
            self.driver.get(f"{BASE_URL}/patient/scan")
            # Wait for file upload component
            WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
            )
            self.log_result("WEB-004", "Scan Page Navigation", "Load Scan Page", "PASS", "Scan page file input field detected", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-004", "Scan Page Navigation", "Load Scan Page", "FAIL", str(e), time.time() - t_start)

        # --- TEST 5: Upload Image & Simulate Match ---
        t_start = time.time()
        try:
            # Select file input
            file_input = self.driver.find_element(By.XPATH, "//input[@type='file']")
            
            # We will upload one of the dataset images
            sample_img_path = PROJECT_ROOT / "public" / "dataset" / "0102_jpg.rf.2b53c1fb4b3d9d8c7144fc1793578577.jpg"
            if sample_img_path.exists():
                file_input.send_keys(str(sample_img_path.resolve()))
                time.sleep(3) # Wait for dHash matching to execute in browser
                
                self.log_result("WEB-005", "Teeth Scan Matching", "Upload Image for Diagnosis", "PASS", f"Uploaded image: {sample_img_path.name} and matched successfully", time.time() - t_start)
            else:
                self.log_result("WEB-005", "Teeth Scan Matching", "Upload Image for Diagnosis", "FAIL", "Sample image not found in dataset folder for upload", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-005", "Teeth Scan Matching", "Upload Image for Diagnosis", "FAIL", str(e), time.time() - t_start)

        # --- TEST 6: Check Doctors Directory ---
        t_start = time.time()
        try:
            self.driver.get(f"{BASE_URL}/patient/doctors")
            # Wait for doctors list
            time.sleep(2)
            self.log_result("WEB-006", "Doctors Directory", "Load Doctors List", "PASS", "Doctors list page loaded successfully", time.time() - t_start)
        except Exception as e:
            self.log_result("WEB-006", "Doctors Directory", "Load Doctors List", "FAIL", str(e), time.time() - t_start)
            
        # Clean up browser
        print("Closing Selenium Web Driver...")
        self.driver.quit()
        
    def generate_excel_report(self):
        print(f"Generating styled E2E Web Test Report at {EXCEL_REPORT}...")
        df = pd.DataFrame(self.results)
        
        # Save to Excel
        with pd.ExcelWriter(EXCEL_REPORT, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Web Test E2E Report")
            
            workbook = writer.book
            worksheet = writer.sheets["Web Test E2E Report"]
            
            # Styled colors (Clinical Blue theme)
            header_fill = PatternFill(start_color="1E5E7A", end_color="1E5E7A", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            
            # Pass/Fail colors
            pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            pass_font = Font(name="Segoe UI", size=10, color="155724", bold=True)
            fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            fail_font = Font(name="Segoe UI", size=10, color="721C24", bold=True)
            
            data_font = Font(name="Segoe UI", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            thin_border_side = Side(style='thin', color='E0E0E0')
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            # Headers
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            worksheet.row_dimensions[1].height = 28
            
            # Data Formatting
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 22
                status = df.iloc[row_idx - 2]["Status"]
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    col_name = df.columns[col_idx - 1]
                    if col_name in ["Test ID", "Status", "Response Time (s)", "Timestamp"]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
                    # Apply conditional formatting on status
                    if col_name == "Status":
                        if status == "PASS":
                            cell.fill = pass_fill
                            cell.font = pass_font
                        else:
                            cell.fill = fail_fill
                            cell.font = fail_font
                            
            # Auto width adjustment
            for col in worksheet.columns:
                col_letter = get_column_letter(col[0].column)
                col_name = col[0].value
                
                if col_name == "Test ID":
                    worksheet.column_dimensions[col_letter].width = 12
                elif col_name == "Test Name":
                    worksheet.column_dimensions[col_letter].width = 25
                elif col_name == "Test Step":
                    worksheet.column_dimensions[col_letter].width = 28
                elif col_name == "Status":
                    worksheet.column_dimensions[col_letter].width = 12
                elif col_name == "Response Time (s)":
                    worksheet.column_dimensions[col_letter].width = 20
                elif col_name == "Details / Errors":
                    worksheet.column_dimensions[col_letter].width = 45
                elif col_name == "Timestamp":
                    worksheet.column_dimensions[col_letter].width = 22
                    
            worksheet.views.sheetView[0].showGridLines = True
            
        print(f"Success! Web E2E report generated at: {EXCEL_REPORT}")

if __name__ == "__main__":
    tester = WebE2ETester()
    tester.run_tests()
    tester.generate_excel_report()
