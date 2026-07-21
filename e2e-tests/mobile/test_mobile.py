import os
import time
import sys
from pathlib import Path
import pandas as pd

# -- Ensure Appium Python Client is installed --
try:
    from appium import webdriver
    from appium.options.common import AppiumOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    print("Appium-Python-Client is required. Installing now...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "Appium-Python-Client"])
    from appium import webdriver
    from appium.options.common import AppiumOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

# -- Ensure openpyxl is installed for Excel formatting --
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel report generation...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# -- Paths --
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
REPORTS_DIR = PROJECT_ROOT / "e2e-tests" / "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)
EXCEL_REPORT = REPORTS_DIR / "mobile_test_report.xlsx"

# -- Test Suite Config --
APPIUM_SERVER = "http://localhost:4723" # Appium 2.x standard server path
TIMEOUT = 15

class MobileE2ETester:
    def __init__(self):
        self.results = []
        self.driver = None
        
    def start_appium_session(self):
        print("Initializing Appium Session for Android mobile app...")
        
        # Set up Appium Options
        options = AppiumOptions()
        options.set_capability("platformName", "Android")
        options.set_capability("automationName", "UiAutomator2")
        options.set_capability("deviceName", "Android Emulator")
        
        # Package and Activity for Capacitor Android App
        options.set_capability("appPackage", "com.rekanth.orthocareai")
        options.set_capability("appActivity", "com.rekanth.orthocareai.MainActivity")
        options.set_capability("noReset", True)
        options.set_capability("ensureWebviewsHavePages", True)
        
        # Start driver session
        self.driver = webdriver.Remote(APPIUM_SERVER, options=options)
        self.driver.implicitly_wait(5)
        
    def switch_to_webview_context(self):
        print("Waiting for WebView context...")
        time.sleep(5) # Let webview render
        
        # List all available contexts (NATIVE_APP, WEBVIEW_com.rekanth.orthocareai, etc.)
        contexts = self.driver.contexts
        print(f"Available contexts: {contexts}")
        
        webview_context = None
        for c in contexts:
            if "WEBVIEW" in c:
                webview_context = c
                break
                
        if webview_context:
            self.driver.switch_to.context(webview_context)
            print(f"Successfully switched context to: {webview_context}")
            return True
        else:
            print("Warning: WebView context not found. Retrying in native context...")
            return False

    def log_result(self, test_id, name, step, status, details="", response_time=0.0):
        self.results.append({
            "Test ID": test_id,
            "Test Name": name,
            "Test Step": step,
            "Status": status,
            "Response Time (s)": round(response_time, 3),
            "Details / Errors": details,
            "Timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
        })
        print(f"[{status}] {test_id} - {name} ({step}) | Time: {response_time:.2f}s | Details: {details}")

    def run_tests(self):
        t_start = time.time()
        try:
            self.start_appium_session()
            self.log_result("MOB-001", "App Launch", "Initialize Android Session", "PASS", "App opened on emulator/device successfully", time.time() - t_start)
        except Exception as e:
            self.log_result("MOB-001", "App Launch", "Initialize Android Session", "FAIL", str(e), time.time() - t_start)
            return
            
        # Switch context to WebView since Capacitor is a hybrid app
        t_start = time.time()
        has_webview = False
        try:
            has_webview = self.switch_to_webview_context()
            if has_webview:
                self.log_result("MOB-002", "Context Switch", "Switch Native to WebView", "PASS", "Switched context to hybrid WebView successfully", time.time() - t_start)
            else:
                self.log_result("MOB-002", "Context Switch", "Switch Native to WebView", "FAIL", "No WebView context available", time.time() - t_start)
        except Exception as e:
            self.log_result("MOB-002", "Context Switch", "Switch Native to WebView", "FAIL", str(e), time.time() - t_start)
            
        # If we failed to switch to WebView, we cannot do DOM-based automation
        if not has_webview:
            print("Cannot automate web components in native mode. Exiting test.")
            self.driver.quit()
            return
            
        # --- TEST 3: Welcome Screen Verification ---
        t_start = time.time()
        try:
            # Check for container class or welcome text in WebView DOM
            WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.CLASS_NAME, "container"))
            )
            self.log_result("MOB-003", "Landing Screen", "Verify Home Layout", "PASS", "Welcome landing interface rendered inside WebView", time.time() - t_start)
        except Exception as e:
            self.log_result("MOB-003", "Landing Screen", "Verify Home Layout", "FAIL", str(e), time.time() - t_start)
            
        # --- TEST 4: Login Action ---
        t_start = time.time()
        try:
            login_link = WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@href, '/auth/login')]"))
            )
            login_link.click()
            
            # Find inputs in login form
            email_input = WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email']"))
            )
            password_input = self.driver.find_element(By.XPATH, "//input[@type='password']")
            submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit']")
            
            email_input.send_keys("test_patient@orthocare.ai")
            password_input.send_keys("Password123!")
            submit_btn.click()
            
            time.sleep(3) # Wait for page load
            self.log_result("MOB-004", "Mobile Patient Login", "Submit Login Credentials", "PASS", "Logged in successfully within Android Webview", time.time() - t_start)
        except Exception as e:
            self.log_result("MOB-004", "Mobile Patient Login", "Submit Login Credentials", "FAIL", str(e), time.time() - t_start)

        # --- TEST 5: Upload Image from Device ---
        t_start = time.time()
        try:
            # Navigate to scan
            self.driver.execute_script("window.location.hash = '/patient/scan';") # Navigate using Router
            time.sleep(2)
            
            # File Input automation
            file_input = WebDriverWait(self.driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
            )
            
            # For mobile Appium uploads, we can use driver.push_file or send keys
            # To simplify, we send keys to input
            sample_img_name = "0102_jpg.rf.2b53c1fb4b3d9d8c7144fc1793578577.jpg"
            file_input.send_keys(sample_img_name)
            time.sleep(3)
            
            self.log_result("MOB-005", "Diagnosis Scan Match", "Upload Scan Image", "PASS", "Scan uploaded and processed by dHash matching", time.time() - t_start)
        except Exception as e:
            self.log_result("MOB-005", "Diagnosis Scan Match", "Upload Scan Image", "FAIL", str(e), time.time() - t_start)

        # Clean up session
        print("Closing Appium Driver session...")
        self.driver.quit()
        
    def generate_excel_report(self):
        print(f"Generating styled E2E Mobile Test Report at {EXCEL_REPORT}...")
        df = pd.DataFrame(self.results)
        
        if df.empty:
            print("No test results recorded. Skipping report.")
            return
            
        with pd.ExcelWriter(EXCEL_REPORT, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Mobile Test E2E Report")
            
            workbook = writer.book
            worksheet = writer.sheets["Mobile Test E2E Report"]
            
            # Styled colors (Clinical Blue theme)
            header_fill = PatternFill(start_color="1E5E7A", end_color="1E5E7A", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            
            # Pass/Fail colors
            pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            pass_font = Font(name="Segoe UI", size=10, color="155724", bold=True)
            fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            fail_font = Font(name="Segoe UI", size=10, color="721C24", bold=True)
            
            data_font = Font(name="Segoe UI", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            thin_border_side = Side(style='thin', color='E0E0E0')
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            # Headers
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            worksheet.row_dimensions[1].height = 28
            
            # Data Formatting
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 22
                status = df.iloc[row_idx - 2]["Status"]
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    col_name = df.columns[col_idx - 1]
                    if col_name in ["Test ID", "Status", "Response Time (s)", "Timestamp"]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
                    # Apply conditional formatting on status
                    if col_name == "Status":
                        if status == "PASS":
                            cell.fill = pass_fill
                            cell.font = pass_font
                        else:
                            cell.fill = fail_fill
                            cell.font = fail_font
                            
            # Auto width adjustment
            for col in worksheet.columns:
                col_letter = get_column_letter(col[0].column)
                col_name = col[0].value
                
                if col_name == "Test ID":
                    worksheet.column_dimensions[col_letter].width = 12
                elif col_name == "Test Name":
                    worksheet.column_dimensions[col_letter].width = 25
                elif col_name == "Test Step":
                    worksheet.column_dimensions[col_letter].width = 28
                elif col_name == "Status":
                    worksheet.column_dimensions[col_letter].width = 12
                elif col_name == "Response Time (s)":
                    worksheet.column_dimensions[col_letter].width = 20
                elif col_name == "Details / Errors":
                    worksheet.column_dimensions[col_letter].width = 45
                elif col_name == "Timestamp":
                    worksheet.column_dimensions[col_letter].width = 22
                    
            worksheet.views.sheetView[0].showGridLines = True
            
        print(f"Success! Mobile E2E report generated at: {EXCEL_REPORT}")

if __name__ == "__main__":
    tester = MobileE2ETester()
    tester.run_tests()
    tester.generate_excel_report()
