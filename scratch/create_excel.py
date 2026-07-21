import os
import json
import sys
import subprocess
from pathlib import Path

# -- Ensure dependencies are installed --
def install_dependencies():
    required_packages = ["pandas", "openpyxl"]
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            print(f"Installing missing dependency: {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

install_dependencies()

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- Paths --
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATASET_DIR = PROJECT_ROOT / "public" / "dataset"
METADATA_FILE = DATASET_DIR / "metadata.json"
EXCEL_OUTPUT = DATASET_DIR / "orthodontic_dataset.xlsx"
CSV_OUTPUT = DATASET_DIR / "orthodontic_dataset.csv"

def main():
    if not METADATA_FILE.exists():
        print(f"Error: {METADATA_FILE} does not exist.")
        sys.exit(1)
        
    print(f"Reading dataset metadata from {METADATA_FILE}...")
    with open(METADATA_FILE, "r", encoding="utf-8") as f:
        records = json.load(f)
        
    print(f"Found {len(records)} entries in metadata.json.")
    
    # Process records
    processed_records = []
    max_suggestions = 0
    
    # First pass to find the maximum number of suggestions
    for r in records:
        sugs = r.get("suggestions", [])
        if isinstance(sugs, list):
            max_suggestions = max(max_suggestions, len(sugs))
            
    print(f"Max suggestions found in a single entry: {max_suggestions}")
    
    for idx, r in enumerate(records, 1):
        filename = r.get("filename", "")
        condition = r.get("condition", "Unknown")
        severity = r.get("severity", "unknown").capitalize()
        score = r.get("score", 0)
        suggestions_list = r.get("suggestions", [])
        
        # Format suggestions as bullet points
        suggestions_text = ""
        if isinstance(suggestions_list, list) and len(suggestions_list) > 0:
            suggestions_text = "\n".join([f"• {s}" for s in suggestions_list])
        elif isinstance(suggestions_list, str):
            suggestions_text = f"• {suggestions_list}"
            suggestions_list = [suggestions_list]
            
        row_dict = {
            "S.No": idx,
            "Image Filename": filename,
            "Condition (Problem)": condition,
            "Severity": severity,
            "Severity Score": score,
            "Suggestions (Solutions)": suggestions_text
        }
        
        # Add separate columns for each suggestion
        for i in range(max_suggestions):
            col_name = f"Suggestion {i+1}"
            if i < len(suggestions_list):
                row_dict[col_name] = suggestions_list[i]
            else:
                row_dict[col_name] = ""
                
        processed_records.append(row_dict)
        
    df = pd.DataFrame(processed_records)
    
    # Save CSV
    print(f"Saving CSV version to {CSV_OUTPUT}...")
    df.to_csv(CSV_OUTPUT, index=False, encoding="utf-8")
    
    # Save styled Excel sheet
    print(f"Creating styled Excel workbook at {EXCEL_OUTPUT}...")
    
    # Create excel writer using openpyxl
    with pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Orthodontic Dataset")
        
        # Get workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["Orthodontic Dataset"]
        
        # Colors (Clinical Teal Theme)
        header_fill = PatternFill(start_color="1E5E7A", end_color="1E5E7A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        # Data Styles
        data_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Borders
        thin_border_side = Side(style='thin', color='E0E0E0')
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Format Headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        worksheet.row_dimensions[1].height = 28
        
        # Format Data Rows
        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 50 # Give vertical room for wrapped text
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                col_name = df.columns[col_idx - 1]
                
                # Alignments & Formatting
                if col_name in ["S.No", "Severity", "Severity Score"]:
                    cell.alignment = center_align
                elif col_name == "Suggestions (Solutions)":
                    cell.alignment = wrap_left_align
                else:
                    cell.alignment = left_align
                    
                # Style Severity Scores dynamically
                if col_name == "Severity Score":
                    cell.font = bold_font
                    
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = col[0].value
            
            # Custom column widths based on expected content
            if col_name == "S.No":
                worksheet.column_dimensions[col_letter].width = 8
            elif col_name == "Image Filename":
                worksheet.column_dimensions[col_letter].width = 38
            elif col_name == "Condition (Problem)":
                worksheet.column_dimensions[col_letter].width = 32
            elif col_name == "Severity":
                worksheet.column_dimensions[col_letter].width = 12
            elif col_name == "Severity Score":
                worksheet.column_dimensions[col_letter].width = 14
            elif col_name == "Suggestions (Solutions)":
                worksheet.column_dimensions[col_letter].width = 50
            elif "Suggestion " in col_name:
                worksheet.column_dimensions[col_letter].width = 35
            else:
                # Default auto width calculation
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        # Enable grid lines explicitly
        worksheet.views.sheetView[0].showGridLines = True
        
    print(f"Success! Excel sheet successfully generated at {EXCEL_OUTPUT}.")

if __name__ == "__main__":
    main()
